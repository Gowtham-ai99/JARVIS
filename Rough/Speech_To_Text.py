import os
import re
import pytesseract
from PIL import Image, ImageEnhance, ImageFilter
from openpyxl import Workbook, load_workbook

# Path to Tesseract OCR executable
pytesseract.pytesseract.tesseract_cmd = r"C:\Program Files\Tesseract-OCR\tesseract.exe"

# Directory containing payment screenshots
image_folder = "C:\\Users\\User\\Downloads\\New folder"

# Excel file to save data
output_file = "WhatsApp_Payment_Data.xlsx"

# Check if Excel file exists, else create one
if not os.path.exists(output_file):
    wb = Workbook()
    sheet = wb.active
    # Create headers
    sheet.append(["Transaction ID", "Amount (Rs)", "Date", "Time", "File Name"])
    wb.save(output_file)

# Load existing workbook
wb = load_workbook(output_file)
sheet = wb.active

# Function to preprocess the image for better OCR results
def preprocess_image(image_path):
    # Open the image
    image = Image.open(image_path)
    
    # Convert to grayscale
    image = image.convert("L")
    
    # Enhance the image (brightness, contrast, sharpness)
    enhancer = ImageEnhance.Contrast(image)
    image = enhancer.enhance(2)
    enhancer = ImageEnhance.Sharpness(image)
    image = enhancer.enhance(2)
    
    # Thresholding to make text clearer
    image = image.point(lambda p: p > 200 and 255)
    
    return image

# Function to extract details using OCR and regex
def extract_details(image_path):
    # Preprocess the image
    image = preprocess_image(image_path)
    
    # Apply OCR to the preprocessed image
    text = pytesseract.image_to_string(image)

    # Print text for debugging (remove this in production)
    print("OCR Text:", text)

    # Extract Transaction ID (ensure it captures IDs like T2411240906356702517376)
    transaction_id = re.search(r"T\d{18}", text)
    transaction_id = transaction_id.group() if transaction_id else "N/A"

    # Extract Amount with Rupee symbol (₹) or just numbers (like ₹21,750 or 899)
    amount = re.search(r"(₹|Rs)?\s?(\d{1,3}(?:,\d{3})*(?:\.\d{1,2})?)", text)  # Flexible pattern for ₹ and Rs
    if amount:
        # Clean the amount to remove commas and extra spaces
        amount = amount.group(2).replace(",", "") if amount.group(2) else "N/A"
    else:
        amount = "N/A"
    
    # Extract Date
    date = re.search(r"\d{1,2} \w{3} \d{4}", text)
    date = date.group() if date else "N/A"

    # Extract Time
    time = re.search(r"\d{1,2}:\d{2} (AM|PM|am|pm)", text)
    time = time.group().upper() if time else "N/A"

    # Extract File Name
    file_name = os.path.basename(image_path)

    return [transaction_id, amount, date, time, file_name]

# Process all images in the folder
for file_name in os.listdir(image_folder):
    if file_name.lower().endswith((".png", ".jpg", ".jpeg")):
        image_path = os.path.join(image_folder, file_name)
        details = extract_details(image_path)
        sheet.append(details)

# Save to Excel
wb.save(output_file)
print("Data saved to", output_file)
